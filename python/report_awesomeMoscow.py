from datetime import datetime, timedelta

d1 = datetime.now()
d2 = d1 + timedelta(days=-14)
import pymongo
from pymongo import MongoClient

client = MongoClient('localhost', 27017)
db = client['AwesomeMoscow']
pipeline = [
    {
        "$match": {
            "_header._created": {"$lte": d1, "$gt": d2}
        }
    },
    {
        "$match": {
            "City.FORMALNAME": {"$ne": "Москва"}
        }
    },
    {
        "$group": {
            "_id": "$City.FORMALNAME",
            "classes_count": {"$sum": 1}
        }
    }
]
classes = list(db.Class.aggregate(pipeline))
pipeline = [
    {
        "$match": {
            "_header._created": {"$lte": d1, "$gt": d2}
        }
    },
    {
        "$lookup": {
            "from": "Class",
            "localField": "CurrentClassId",
            "foreignField": "_id",
            "as": "Class"
        }
    },
    {
        "$unwind": {
            "path": "$Class",
            "includeArrayIndex": "arrayIndex",
            "preserveNullAndEmptyArrays": True
        }
    },
    {
        "$match": {
            "Class.City.FORMALNAME": {"$ne": "Москва"}
        }
    },
    {
        "$group": {
            "_id": "$Class.City.FORMALNAME",
            "users_count": {"$sum": 1}
        }
    }
]
users = list(db.Profile.aggregate(pipeline))
pipeline = [
    {
        "$match": {
            "_header._created": {"$lte": d1, "$gt": d2}
        }
    },
    {
        "$lookup": {
            "from": "Class",
            "localField": "ClassId",
            "foreignField": "_id",
            "as": "Class"
        }
    },
    {
        "$unwind": {
            "path": "$Class",
            "includeArrayIndex": "arrayIndex",
            "preserveNullAndEmptyArrays": True
        }
    },
    {
        "$match": {
            "Class.City.FORMALNAME": {"$ne": "Москва"}
        }
    },
    {
        "$group": {
            "_id": "$Class.City.FORMALNAME",
            "votes_count": {"$sum": 1}
        }
    }
]
votes = list(db.Vote.aggregate(pipeline))
pipeline = [
    {
        "$match": {
            "_header._created": {"$lte": d1, "$gt": d2}
        }
    },
    {
        "$lookup": {
            "from": "Class",
            "localField": "ClassId",
            "foreignField": "_id",
            "as": "Class"
        }
    },
    {
        "$unwind": {
            "path": "$Class",
            "includeArrayIndex": "arrayIndex",
            "preserveNullAndEmptyArrays": True
        }
    },
    {
        "$match": {
            "Class.City.FORMALNAME": {"$ne": "Москва"}
        }
    },
    {
        "$group": {
            "_id": "$Class.City.FORMALNAME",
            "events_count": {"$sum": 1}
        }
    }
]
events = list(db.Event.aggregate(pipeline))
#join classes and events
for x in classes:
    r = next((item for item in users if item["_id"] == x['_id']), None)
    x['users_count'] = r['users_count'] if (r) else 0
    r = next((item for item in votes if item["_id"] == x['_id']), None)
    x['votes_count'] = r['votes_count'] if (r) else 0
    r = next((item for item in events if item["_id"] == x['_id']), None)
    x['events_count'] = r['events_count'] if (r) else 0

pipeline = [
    {
        "$match": {
            "_header._created": {"$lte": d1, "$gt": d2}
        }
    },
    {
        "$lookup": {
            "from": "Class",
            "localField": "ClassId",
            "foreignField": "_id",
            "as": "embeddedData"
        }
    },
    {
        "$unwind": "$embeddedData"
    },
    {
        "$project": {
            "Название_Класса": "$embeddedData.ClassName",
            "Название_События": "$Name",
            "Место": {"$ifNull": ["$Place", "-"]},
            "Дата_создания": "$_header._created",
            "Создано_на_основе_идеи": {
                "$cond": [{"$not": "$SourceIdeaUrl"}, "Нет", "Да"]
            },
            "Идея": "$SourceIdeaUrl",
            "_id": 0,
            "Город": "$embeddedData.City.FORMALNAME",
            "Количество_участников": {
                "$size": "$Participants"
            }
        }
    },
    {
        "$match": {
            "Город": "Москва"
        }
    }]
res = list(db.Event.aggregate(pipeline))
from openpyxl import Workbook
from openpyxl import load_workbook

import os
dir_path = os.path.dirname(os.path.realpath(__file__))

wb = load_workbook(dir_path + '/template_moscow.xlsx')
wb2 =  load_workbook(dir_path + '/template_zamkad.xlsx')
ws2 = wb2.active
k = 2
for x in classes:
    ws2.cell(row=k, column=1, value=x['_id'])
    ws2.cell(row=k, column=2, value=x['users_count'])
    ws2.cell(row=k, column=3, value=x['classes_count'])
    ws2.cell(row=k, column=4, value=x['votes_count'])
    ws2.cell(row=k, column=5, value=x['events_count'])
    k += 1

ws = wb['События']
k = 2
for x in res:
    ws.cell(row=k, column=1, value=x['Название_Класса'])
    ws.cell(row=k, column=2, value=x['Название_События'])
    ws.cell(row=k, column=3, value=x['Место'])
    ws.cell(row=k, column=4, value=x['Дата_создания'])
    ws.cell(row=k, column=5, value=x['Создано_на_основе_идеи'])
    ws.cell(row=k, column=6, value=x['Количество_участников'])
    k += 1
pipeline = [
    {
        "$match": {
            "_header._created": {"$lte": d1, "$gt": d2}
        }
    },
    {
        "$lookup": {
            "from": "Class",
            "localField": "ClassId",
            "foreignField": "_id",
            "as": "embeddedData"
        }
    },
    {
        "$unwind": {
            "path": "$embeddedData",
            "preserveNullAndEmptyArrays": True
        }
    },
    {
        "$project": {
            "Название_Класса": {"$ifNull": ["$embeddedData.ClassName", "-"]},
            "Название_Голосования": "$VoteName",
            "Дата_создания": "$_header._created",
            "Создано_на_основе_идеи": {
                "$cond": [{"$not": "$SourceIdeaUrl"}, "Нет", "Да"]
            },
            "_id": 0,
            "Город": "$embeddedData.City.FORMALNAME",
            "Количество_голосов": {"$size": {"$ifNull": ["$Users", []]}}
        }
    },
    {
        "$match": {
            "Город": "Москва"
        }
    }
]
res = list(db.Vote.aggregate(pipeline))
ws = wb['Голосования']
k = 2
for x in res:
    ws.cell(row=k, column=1, value=x['Название_Класса'])
    ws.cell(row=k, column=2, value=x['Название_Голосования'])
    ws.cell(row=k, column=3, value=x['Дата_создания'])
    ws.cell(row=k, column=4, value=x['Создано_на_основе_идеи'])
    ws.cell(row=k, column=5, value=x['Количество_голосов'])
    k += 1
pipeline = [
    {
        "$match": {
            "_header._created": {"$lte": d1, "$gt": d2}
        }
    },
    {
        "$lookup": {
            "from": "Profile",
            "localField": "_id",
            "foreignField": "CurrentClassId",
            "as": "embeddedData"

        }
    },
    {
        "$project": {
            "_id": 0,
            "Название_Класса": {"$ifNull": ["$ClassName", "-"]},
            "Дата_создания": "$_header._created",
            "Город": "$City.FORMALNAME",
            "Количество_Родителей_в_классе": {"$size": {"$ifNull": ["$embeddedData", []]}}
        }
    },
    {
        "$match": {
            "Город": "Москва"
        }
    }
]
res = list(db.Class.aggregate(pipeline))
ws = wb['Классы']
k = 2
for x in res:
    ws.cell(row=k, column=1, value=x['Название_Класса'])
    ws.cell(row=k, column=2, value=x['Дата_создания'])
    ws.cell(row=k, column=3, value=x['Количество_Родителей_в_классе'])
    k += 1
pipeline = [
    {
        "$match": {
            "_header._created": {"$lte": d1, "$gt": d2}
        }
    },
    {
        "$lookup": {
            "from": "Profile",
            "localField": "_id",
            "foreignField": "CurrentClassId",
            "as": "embeddedData"
        }
    },
    {
        "$project": {
            "_id": 0,
            "Название_Класса": {"$ifNull": ["$ClassName", "-"]},
            "Дата_создания": "$_header._created",
            "Город": "$City.FORMALNAME",
            "Количество_Родителей_в_классе": {"$size": {"$ifNull": ["$embeddedData", []]}}
        }
    },
    {
        "$match": {
            "Город": "Москва"
        }
    }
]
res = list(db.Class.aggregate(pipeline))
ws = wb['Классы']
k = 2
for x in res:
    ws.cell(row=k, column=1, value=x['Название_Класса'])
    ws.cell(row=k, column=2, value=x['Дата_создания'])
    ws.cell(row=k, column=3, value=x['Количество_Родителей_в_классе'])
    k += 1
pipeline = [
    {
        "$match": {
            "_header._created": {"$lte": d1, "$gt": d2}
        }
    },
    {
        "$unwind": {
            "path": "$Roles",
            "preserveNullAndEmptyArrays": True
        }
    },
    {
        "$lookup": {
            "from": "Class",
            "localField": "Roles.ClassId",
            "foreignField": "_id",
            "as": "embeddedData"

        }
    },
    {
        "$unwind": {
            "path": "$embeddedData",
            "preserveNullAndEmptyArrays": True
        }
    },
    {
        "$match": {
            "embeddedData.City.FORMALNAME": "Москва"
        }
    },
    {
        "$group": {
            "_id": "$FullName",
            "count": {"$sum": int(1)},
            "Создан": {"$first": "$_header._created"}
        }
    },
    {
        "$project": {
            "Имя пользователя": "$_id",
            "Дата регистрации": "$Создан",
            "Количество классов": "$count",
            "_id": 0
        }
    }
]
res = list(db.Profile.aggregate(pipeline))
ws = wb['Пользователи']
k = 2
for x in res:
    ws.cell(row=k, column=1, value=x['Имя пользователя'])
    ws.cell(row=k, column=2, value=x['Дата регистрации'])
    ws.cell(row=k, column=3, value=x['Количество классов'])
    k += 1

filename_moscow = "Moscow " + d2.strftime("%Y-%m-%d") + ' - ' + d1.strftime("%Y-%m-%d") + ".xlsx"
filename_zamkad = "Other cities " + d2.strftime("%Y-%m-%d") + ' - ' + d1.strftime("%Y-%m-%d") + ".xlsx"
wb.save(dir_path + "/" +  filename_moscow)
wb2.save(dir_path + "/" + filename_zamkad)

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

server = smtplib.SMTP_SSL('smtp.yandex.ru:465')
login = "noreply@xn--80aa2aelxa8i.xn--80adxhks"
passw = "5WD8TXemTDuXdGyv"
server.login(login, passw)

msg = MIMEMultipart()
recipients = ['i.ivanov@x.ru', 'i.ivanov@yandex.ru']
msg['From'] = login
msg['To'] = ', '.join(recipients)
msg['Subject'] = 'Отчет об активностях в сервисе Классная.Москва за ' + d2.strftime("%Y-%m-%d") + '..' + d1.strftime("%Y-%m-%d")

part = MIMEBase('application', "octet-stream")
part.set_payload(open(dir_path + "/" +  filename_moscow, "rb").read())
encoders.encode_base64(part)

part.add_header('Content-Disposition', 'attachment; filename="' + filename_moscow + '"')
msg.attach(part)

part = MIMEBase('application', "octet-stream")
part.set_payload(open(dir_path + "/" +  filename_zamkad, "rb").read())
encoders.encode_base64(part)

part.add_header('Content-Disposition', 'attachment; filename="' + filename_zamkad + '"')

msg.attach(part)

server.sendmail(login, recipients, msg.as_string())
server.quit()